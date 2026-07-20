import json
import re
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Optional, Tuple

from openpyxl import load_workbook
from sqlalchemy.orm import Session

try:
    from .models import Booking, DatasetMeta, TeamMapping
except ImportError:  # pragma: no cover
    from models import Booking, DatasetMeta, TeamMapping


REQUIRED_SBTR_COLUMNS = [
    "Booked Date",
    "Project Name",
    "Unit No.",
    "Sales Head",
    "Attended by",
    "Team Head",
    "Sr. VP",
    "CSO",
    "Sale Value (AED)",
    "PNS",
    "Unit Count",
    "Collection 10%",
]


def excel_date_to_iso(serial: Any) -> Optional[str]:
    if serial is None or serial == "":
        return None
    if isinstance(serial, datetime):
        return serial.date().isoformat()
    try:
        n = float(serial)
    except (TypeError, ValueError):
        return None
    if n != n:  # NaN
        return None
    base = datetime(1899, 12, 30)
    return (base + timedelta(days=n)).date().isoformat()


def region_from_project(project: Optional[str]) -> str:
    if not project:
        return "Dubai"
    if project in ("Sobha Siniya Island", "Downtown UAQ"):
        return "UAQ"
    if project == "Sobha City":
        return "Abu Dhabi"
    return "Dubai"


def map_sbtr_row(row: dict[str, Any], full_sv_key: Optional[str]) -> dict[str, Any]:
    booked = excel_date_to_iso(row.get("Booked Date"))
    cancelled = str(row.get("Cancelled Units") or "").strip().lower() == "yes"
    collection10 = str(row.get("Collection 10%") or "").strip().lower() == "yes"
    if cancelled:
        status = "C"
    elif collection10:
        status = "Q"
    else:
        status = "N"

    sales_head = row.get("Sales Head") or ""
    attended = row.get("Attended by") or ""
    is_london = bool(re.match(r"^Gopeshwar Mahato$", str(sales_head), re.I)) or (
        sales_head == "Vinit Manishbhai Parikh" and attended != "Vinit Manishbhai Parikh"
    )
    team_head = row.get("Team Head") or ""
    sr_vp = row.get("Sr. VP") or ""
    cso = row.get("CSO") or ""
    vp = "London Team" if is_london else (team_head or sr_vp or cso or "Unassigned")

    raw_sv = row.get(full_sv_key) if full_sv_key and row.get(full_sv_key) is not None else row.get("Sale Value (AED)")
    try:
        sv = float(raw_sv or 0)
    except (TypeError, ValueError):
        sv = 0.0
    try:
        pns = float(row.get("PNS") or 0)
    except (TypeError, ValueError):
        pns = 0.0
    try:
        uc = float(row.get("Unit Count") or 0)
    except (TypeError, ValueError):
        uc = 0.0
    try:
        pp = round(float(row.get("Paid %") or 0) * 1000) / 10
    except (TypeError, ValueError):
        pp = 0.0

    return {
        "d": booked,
        "m": booked[:7] if booked else "",
        "p": row.get("Project Name") or "",
        "l": region_from_project(row.get("Project Name")),
        "sm": attended,
        "sd": "London Team" if is_london else sales_head,
        "vp": vp,
        "cso": cso,
        "r_sm": attended,
        "r_sd": sales_head,
        "r_th": team_head,
        "r_srvp": sr_vp,
        "r_cso": cso,
        "sv": sv,
        "pns": pns,
        "uc": uc,
        "s": status,
        "u": row.get("Unit No.") or "",
        "ut": row.get("No. of Bedrooms") or "",
        "pp": pp,
        "br": row.get("Broker Company Name") or "Direct",
        "d10": excel_date_to_iso(row.get("10%_Date")),
        "d20": excel_date_to_iso(row.get("20%_Date")),
        "spa": row.get("SPA Executed") or "",
        "dld": row.get("DLD Status") or "",
    }


def _sheet_rows(file_bytes: bytes) -> Tuple[list[str], list[dict[str, Any]]]:
    wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if not header_row:
        raise ValueError("The sheet is empty.")
    headers = [
        (str(h).strip() if h is not None and str(h).strip() else f"Column {i + 1}")
        for i, h in enumerate(header_row)
    ]
    records: list[dict[str, Any]] = []
    for row in rows_iter:
        if not row or not any(c not in (None, "") for c in row):
            continue
        item = {}
        for i, h in enumerate(headers):
            val = row[i] if i < len(row) else None
            if isinstance(val, datetime):
                item[h] = val
            elif val is None:
                item[h] = ""
            else:
                item[h] = val
        records.append(item)
    wb.close()
    return headers, records


def parse_sbtr_excel(file_bytes: bytes) -> Tuple[list[dict[str, Any]], str]:
    headers, rows = _sheet_rows(file_bytes)
    if not rows:
        raise ValueError("No data rows found.")

    sample = rows[0]
    missing = [col for col in REQUIRED_SBTR_COLUMNS if col not in sample]
    if missing:
        raise ValueError(f"This file is missing required SBTR columns: {', '.join(missing)}")

    full_sv_key = next(
        (h for h in headers if h and h.strip().lower() == "full sale value"),
        headers[-2] if len(headers) >= 2 else None,
    )

    records = [map_sbtr_row(r, full_sv_key) for r in rows]
    records = [r for r in records if r.get("d") and r.get("u")]
    if not records:
        raise ValueError("No valid rows after parsing.")

    dates = sorted(r["d"] for r in records if r.get("d"))
    latest = dates[-1]
    d = datetime.strptime(latest, "%Y-%m-%d")
    as_of = f"{d.day} {d.strftime('%b %Y')}"
    return records, as_of


def parse_team_mapping_excel(file_bytes: bytes, file_name: str) -> dict[str, Any]:
    headers, rows = _sheet_rows(file_bytes)
    if not rows:
        raise ValueError("No data rows found in team mapping file.")
    table_rows = []
    for r in rows:
        table_rows.append([str(r.get(h, "") if r.get(h) is not None else "") for h in headers])
    return {
        "fileName": file_name,
        "headers": headers,
        "rows": table_rows,
    }


def replace_bookings(db: Session, records: list[dict[str, Any]], as_of: str) -> int:
    db.query(Booking).delete()
    db.bulk_save_objects([Booking(data=json.dumps(r)) for r in records])
    meta = db.query(DatasetMeta).order_by(DatasetMeta.id.desc()).first()
    if not meta:
        meta = DatasetMeta()
        db.add(meta)
    meta.sbtr_as_of = as_of
    meta.row_count = len(records)
    meta.updated_at = datetime.utcnow()
    db.commit()
    return len(records)


def replace_team_mapping(db: Session, payload: dict[str, Any]) -> dict[str, Any]:
    db.query(TeamMapping).delete()
    entry = TeamMapping(
        file_name=payload.get("fileName") or "team_mapping.xlsx",
        payload=json.dumps(payload),
    )
    db.add(entry)
    # Bump dataset meta so open dashboards detect the change via /meta polling.
    meta = db.query(DatasetMeta).order_by(DatasetMeta.id.desc()).first()
    if not meta:
        meta = DatasetMeta()
        db.add(meta)
    meta.updated_at = datetime.utcnow()
    db.commit()
    return payload


def get_bookings(db: Session) -> list[dict[str, Any]]:
    return [json.loads(row.data) for row in db.query(Booking).all()]


def get_team_mapping(db: Session) -> Optional[dict[str, Any]]:
    entry = db.query(TeamMapping).order_by(TeamMapping.id.desc()).first()
    if not entry:
        return None
    return json.loads(entry.payload)


def get_meta(db: Session) -> dict[str, Any]:
    meta = db.query(DatasetMeta).order_by(DatasetMeta.id.desc()).first()
    if not meta:
        return {"sbtr_as_of": None, "row_count": 0}
    return {
        "sbtr_as_of": meta.sbtr_as_of,
        "row_count": meta.row_count,
        "updated_at": meta.updated_at.isoformat() if meta.updated_at else None,
    }
